import sqlite3, pandas as pd, qrcode, os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

db_path = "instance/stock.db"
base_url = "gabrielsnewpc.yak-mohs.ts.net"

try:
    test_conn = sqlite3.connect(db_path)
    test_conn.execute("SELECT 1")
    test_conn.close()
except sqlite3.OperationalError as e:
    print("Error: Could not access the database.")
    print("Details:", e)
    exit()

conn = sqlite3.connect(db_path)
item = pd.read_sql("SELECT * FROM item", conn)
cat = pd.read_sql("SELECT * FROM category", conn)
conn.close()

if not base_url.startswith("http://") and not base_url.startswith("https://"):
    base_url = "https://" + base_url
if not base_url.endswith("/"):
    base_url += "/"

item["qr_link"] = base_url + "edit/" + item["uid"].astype(str) + "cat=" + item["category_id"].fillna(0).astype(int).astype(str)

df = item.merge(cat, how="left", left_on="category_id", right_on="id")
df.rename(columns={"name_x": "item_name", "name_y": "category_name"}, inplace=True)
df["category_name"] = df["category_name"].fillna("Uncategorized").astype(str)

os.makedirs("qr", exist_ok=True)
for f in os.listdir("qr"):
    os.remove(os.path.join("qr", f))

df["qr_path"] = [f"qr/qr_{i}.png" for i in range(len(df))]
for i, row in df.iterrows():
    qrcode.make(row["qr_link"]).save(row["qr_path"])

wb = Workbook()
ws = wb.active
ws.title = "Items with QR Codes"

text_columns = ["item_name", "count", "uid"]
for i, col in enumerate(text_columns, start=1):
    width = df[col].astype(str).map(len).max()
    pad = 0.5 if col == "item_name" else 2
    ws.column_dimensions[get_column_letter(i)].width = width + pad
ws.column_dimensions['D'].width = len(base_url) + 2
ws.column_dimensions['E'].width = 15

df.sort_values("category_id", inplace=True)
r = 1
for cid, g in df.groupby("category_id"):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(r, 1, f"Category #{int(cid)}")
    r += 1
    for c, h in enumerate(["Item Name", "Count", "UID", "URL", "QR Code"], 1):
        ws.cell(r, c, h).alignment = Alignment(horizontal="center", vertical="center")
    r += 1
    for _, row in g.iterrows():
        ws.cell(r, 1, row["item_name"])
        ws.cell(r, 2, row["count"])
        ws.cell(r, 3, row["uid"])
        ws.cell(r, 4, base_url)
        img = ExcelImage(row["qr_path"])
        img.width = 100
        img.height = 100
        ws.row_dimensions[r].height = 75
        ws.add_image(img, f"E{r}")
        r += 1

for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

wb.save("stock.xlsx")
